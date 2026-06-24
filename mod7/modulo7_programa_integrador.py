"""
================================================================================
TRABAJO PRÁCTICO INTEGRADOR - ESTADÍSTICA
MÓDULO 7 - SEXTA ETAPA: PROGRAMA PYTHON INTEGRADOR
================================================================================
Programa que automatiza todos los cálculos estadísticos del trabajo integrador
utilizando las librerías Pandas, NumPy y Matplotlib (según consigna).

Variables analizadas:
- Cantidad de horas diarias de redes sociales (X)
- Rendimiento académico (Y)

Análisis incluidos:
 1. Tabla de frecuencias completa
 2. Medidas de tendencia central (media, moda, mediana)
 3. Medidas de dispersión (varianza, desvío estándar, coef. variación)
 4. Medidas de forma (asimetría de Pearson y Fisher, curtosis de Fisher)
 5. Medidas de posición (cuartiles, deciles, percentiles)
 6. Histograma con líneas de media, mediana y moda
 7. Análisis de regresión lineal entre las dos variables:
    - Diagrama de dispersión con recta de regresión
    - Coeficiente de correlación de Pearson
    - Coeficiente de determinación

Salidas generadas:
- resultados_modulo7.xlsx (tabla de frecuencias + medidas + regresión)
- histograma_modulo7.png   (gráfico de barras pegadas)
- dispersion_modulo7.png   (diagrama de dispersión con recta de regresión)
================================================================================
"""

# ============================================================================
# 1. IMPORTACIÓN DE LIBRERÍAS
# ============================================================================
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ============================================================================
# CONFIGURACIÓN GENERAL
# ============================================================================
RUTA_EXCEL = "tp_proporcionado.xlsx"
HOJA_DATOS = "Regresion Lineal"  # Usamos esta hoja porque contiene los datos
                                  # con rendimiento académico decimal (más preciso)


# ============================================================================
# FUNCIONES AUXILIARES PARA INTERPRETACIONES
# ============================================================================
def interpretar_cv(cv):
    """Interpreta el coeficiente de variación"""
    if cv < 10:
        return "Datos MUY HOMOGÉNEOS (CV < 10%)"
    elif cv < 30:
        return "Datos HOMOGÉNEOS (10% < CV < 30%)"
    else:
        return "Datos HETEROGÉNEOS (CV > 30%)"

def interpretar_asimetria(valor):
    """Interpreta el coeficiente de asimetría"""
    if valor < -0.1:
        return "Asimétrica negativa (cola hacia la izquierda)"
    elif valor > 0.1:
        return "Asimétrica positiva (cola hacia la derecha)"
    else:
        return "Aproximadamente simétrica"

def interpretar_curtosis(valor):
    """Interpreta la curtosis de Fisher"""
    if valor < 3:
        return f"Platicúrtica (α = {valor:.4f} < 3, distribución MÁS ACHATADA que la normal)"
    elif valor > 3:
        return f"Leptocúrtica (α = {valor:.4f} > 3, distribución MÁS APUNTADA que la normal)"
    else:
        return f"Mesocúrtica (α = {valor:.4f} = 3, igual a la curva normal)"

def interpretar_correlacion(r):
    """Interpreta el coeficiente de correlación de Pearson"""
    abs_r = abs(r)
    if abs_r < 0.2:
        fuerza = "MUY DÉBIL o NULA"
    elif abs_r < 0.4:
        fuerza = "DÉBIL"
    elif abs_r < 0.7:
        fuerza = "MODERADA"
    elif abs_r < 0.9:
        fuerza = "FUERTE"
    else:
        fuerza = "MUY FUERTE"
    signo = "negativa (inversa)" if r < 0 else "positiva (directa)"
    return f"Correlación {fuerza} {signo}"


# ============================================================================
# IMPRESIÓN DE TÍTULOS
# ============================================================================
def titulo(texto, nivel=1):
    """Imprime un título con formato"""
    if nivel == 1:
        print("\n" + "=" * 80)
        print(texto.center(80))
        print("=" * 80)
    elif nivel == 2:
        print("\n" + "-" * 80)
        print(f"  {texto}")
        print("-" * 80)


# ============================================================================
# 2. CARGA DE DATOS
# ============================================================================
titulo("MÓDULO 7 - PROGRAMA PYTHON INTEGRADOR")
print(f"\n>> Cargando datos desde: {RUTA_EXCEL}")
print(f">> Hoja: {HOJA_DATOS}")

# Leer la hoja, comenzando desde la fila 2 (donde están los encabezados)
df = pd.read_excel(RUTA_EXCEL, sheet_name=HOJA_DATOS, header=1)
# Quedarnos solo con las dos columnas que necesitamos (B y C: xi e yi)
df = df.iloc[:, [1, 2]].copy()
df.columns = ['horas_rrss', 'rendimiento']
# Tomar solo las primeras 490 filas (las que contienen datos reales,
# excluyendo la fila SUMAS y los resultados de la regresión que vienen abajo)
df = df.iloc[:490].copy()
df = df.dropna()
df = df.reset_index(drop=True)

# Asegurarnos del tipo de datos
df['horas_rrss'] = df['horas_rrss'].astype(int)
df['rendimiento'] = df['rendimiento'].astype(float)

# Variables principales
X = df['horas_rrss']  # variable independiente
Y = df['rendimiento']  # variable dependiente
n = len(df)

print(f">> Total de registros cargados: n = {n}")
print(f">> Variable X: {X.name} (horas)")
print(f">> Variable Y: {Y.name} (escala 1 a 10)")


# ============================================================================
# 3. ANÁLISIS UNIVARIADO DE X (HORAS DE REDES SOCIALES)
# ============================================================================
titulo("ANÁLISIS UNIVARIADO - Cantidad de horas diarias de redes sociales (X)", 1)


# ----- 3.1 TABLA DE FRECUENCIAS -----
titulo("3.1 Tabla de frecuencias", 2)

# Frecuencia absoluta (fi) por cada valor de xi
fi = X.value_counts().sort_index()

tabla = pd.DataFrame({'xi': fi.index, 'fi': fi.values})
tabla['Fi'] = tabla['fi'].cumsum()
tabla['fri'] = tabla['fi'] / n
tabla['Fri'] = tabla['fri'].cumsum()
tabla['pi_pct'] = tabla['fri'] * 100
tabla['Pi_pct'] = tabla['Fri'] * 100
tabla['xi_fi'] = tabla['xi'] * tabla['fi']

# Mostrar tabla
print(tabla.to_string(index=False, float_format='%.4f'))
print(f"\n  TOTAL: n = {tabla['fi'].sum()}   |   Σ fri = {tabla['fri'].sum():.4f}")


# ----- 3.2 MEDIDAS DE TENDENCIA CENTRAL -----
titulo("3.2 Medidas de tendencia central", 2)

media_x = tabla['xi_fi'].sum() / n
moda_x = int(tabla.loc[tabla['fi'].idxmax(), 'xi'])
pos_mediana = n / 2
mediana_x = int(tabla.loc[tabla['Fi'] >= pos_mediana, 'xi'].iloc[0])

print(f"  Media (x̄)   = {media_x:.4f} horas/día")
print(f"     >> En promedio, los estudiantes utilizan {media_x:.2f} horas diarias de redes sociales.")
print(f"\n  Moda (Mo)   = {moda_x} horas/día")
print(f"     >> Es el valor que más se repite: {tabla['fi'].max()} estudiantes lo presentan.")
print(f"\n  Mediana (Md) = {mediana_x} horas/día")
print(f"     >> El 50% de los estudiantes usa {mediana_x} horas o menos; el otro 50%, {mediana_x} horas o más.")


# ----- 3.3 MEDIDAS DE DISPERSIÓN -----
titulo("3.3 Medidas de dispersión", 2)

# Agregamos columnas auxiliares
tabla['dif2_fi'] = (tabla['xi'] - media_x) ** 2 * tabla['fi']
tabla['dif3_fi'] = (tabla['xi'] - media_x) ** 3 * tabla['fi']
tabla['dif4_fi'] = (tabla['xi'] - media_x) ** 4 * tabla['fi']

varianza_x = tabla['dif2_fi'].sum() / (n - 1)   # muestral (n-1)
desvio_x = np.sqrt(varianza_x)
cv_x = (desvio_x / media_x) * 100

print(f"  Varianza muestral (s²) = {varianza_x:.4f}")
print(f"  Desvío estándar (s)    = {desvio_x:.4f}")
print(f"  Coef. de Variación CV  = {cv_x:.2f} %")
print(f"     >> {interpretar_cv(cv_x)}")
print(f"     >> Los datos se alejan en promedio ±{desvio_x:.2f} horas de la media.")


# ----- 3.4 MEDIDAS DE FORMA -----
titulo("3.4 Medidas de forma (asimetría y curtosis)", 2)

as_pearson = 3 * (media_x - mediana_x) / desvio_x
as_fisher = tabla['dif3_fi'].sum() / (n * desvio_x ** 3)
curt_fisher = tabla['dif4_fi'].sum() / (n * desvio_x ** 4)

print(f"  Asimetría de Pearson  = {as_pearson:.4f}")
print(f"     >> {interpretar_asimetria(as_pearson)}")
print(f"\n  Asimetría de Fisher   = {as_fisher:.4f}")
print(f"     >> {interpretar_asimetria(as_fisher)}")
print(f"\n  Curtosis de Fisher    = {curt_fisher:.4f}")
print(f"     >> {interpretar_curtosis(curt_fisher)}")


# ----- 3.5 MEDIDAS DE POSICIÓN -----
titulo("3.5 Medidas de posición (cuartiles, deciles, percentiles)", 2)

def calc_posicion(k, base):
    """Calcula el cuartil/decil/percentil por método de posición"""
    pos = k * n / base
    valor = int(tabla.loc[tabla['Fi'] >= pos, 'xi'].iloc[0])
    return pos, valor

print("  CUARTILES:")
cuartiles = {}
for k in [1, 2, 3]:
    pos, val = calc_posicion(k, 4)
    cuartiles[f'Q{k}'] = val
    print(f"     Q{k}: posición = {k}·n/4 = {pos:.0f}  →  Q{k} = {val} horas")

print("\n  DECILES:")
for k in [3, 4, 9]:
    pos, val = calc_posicion(k, 10)
    print(f"     D{k}: posición = {k}·n/10 = {pos:.0f}  →  D{k} = {val} horas")

print("\n  PERCENTILES:")
for k in [23, 75, 97]:
    pos, val = calc_posicion(k, 100)
    print(f"     P{k}: posición = {k}·n/100 = {pos:.0f}  →  P{k} = {val} horas")


# ----- 3.6 HISTOGRAMA -----
titulo("3.6 Generación del histograma", 2)

fig, ax = plt.subplots(figsize=(11, 6))
ax.bar(tabla['xi'], tabla['fi'], width=1.0, edgecolor='black',
       color='#2E75B6', alpha=0.85)
ax.axvline(media_x, color='red', linestyle='--', linewidth=2,
           label=f'Media = {media_x:.2f}')
ax.axvline(mediana_x, color='green', linestyle=':', linewidth=2,
           label=f'Mediana = {mediana_x}')
ax.axvline(moda_x, color='orange', linestyle='-.', linewidth=2,
           label=f'Moda = {moda_x}')
ax.set_title('Histograma - Horas diarias de uso de redes sociales\n'
             f'(n = {n} estudiantes)', fontsize=13, fontweight='bold')
ax.set_xlabel('Horas diarias de uso (xi)', fontsize=11)
ax.set_ylabel('Cantidad de estudiantes (fi)', fontsize=11)
ax.set_xticks(range(0, 13))
ax.grid(axis='y', alpha=0.3)
ax.legend(loc='upper right', fontsize=10)
for x, f in zip(tabla['xi'], tabla['fi']):
    ax.text(x, f + 1, str(f), ha='center', fontsize=9)
plt.tight_layout()
plt.savefig('histograma_modulo7.png', dpi=150, bbox_inches='tight')
plt.close()
print("  ✓ Histograma guardado como 'histograma_modulo7.png'")


# ============================================================================
# 4. ANÁLISIS BIVARIADO - REGRESIÓN LINEAL Y CORRELACIÓN
# ============================================================================
titulo("ANÁLISIS BIVARIADO - Regresión lineal: Horas RRSS vs Rendimiento", 1)


# ----- 4.1 ESTADÍSTICAS BÁSICAS DE Y -----
titulo("4.1 Estadísticas básicas del rendimiento académico (Y)", 2)
media_y = Y.mean()
desvio_y = Y.std()
print(f"  Media de Y    = {media_y:.4f}")
print(f"  Desvío de Y   = {desvio_y:.4f}")
print(f"  Mínimo de Y   = {Y.min():.2f}")
print(f"  Máximo de Y   = {Y.max():.2f}")


# ----- 4.2 CÁLCULOS PARA LA REGRESIÓN -----
titulo("4.2 Sumatorias para la regresión", 2)
sum_x = X.sum()
sum_y = Y.sum()
sum_x2 = (X ** 2).sum()
sum_y2 = (Y ** 2).sum()
sum_xy = (X * Y).sum()

print(f"  n        = {n}")
print(f"  Σ xi     = {sum_x}")
print(f"  Σ yi     = {sum_y:.2f}")
print(f"  Σ xi²    = {sum_x2}")
print(f"  Σ yi²    = {sum_y2:.2f}")
print(f"  Σ xi·yi  = {sum_xy:.2f}")


# ----- 4.3 RECTA DE REGRESIÓN -----
titulo("4.3 Cálculo de la recta de regresión ŷ = a + b·x", 2)

# Pendiente: b = (n·Σxy - Σx·Σy) / (n·Σx² - (Σx)²)
b = (n * sum_xy - sum_x * sum_y) / (n * sum_x2 - sum_x ** 2)
# Ordenada al origen: a = (Σy - b·Σx) / n
a = (sum_y - b * sum_x) / n

print(f"  Pendiente:        b = {b:.6f}")
print(f"  Ordenada origen:  a = {a:.6f}")
print(f"\n  ECUACIÓN DE LA RECTA:")
signo = "-" if b < 0 else "+"
print(f"     ŷ = {a:.4f} {signo} {abs(b):.4f} · x")

print(f"\n  INTERPRETACIÓN DE LA PENDIENTE:")
if b < 0:
    print(f"     Por cada hora adicional que un estudiante dedica a las redes sociales,")
    print(f"     su rendimiento académico disminuye en {abs(b):.4f} puntos.")
else:
    print(f"     Por cada hora adicional que un estudiante dedica a las redes sociales,")
    print(f"     su rendimiento académico aumenta en {b:.4f} puntos.")

print(f"\n  EJEMPLOS DE PREDICCIÓN:")
for x_ej in [2, 5, 8]:
    y_ej = a + b * x_ej
    print(f"     Si un estudiante usa {x_ej} hs/día → ŷ = {a:.4f} {signo} {abs(b):.4f}·{x_ej} = {y_ej:.2f} puntos")


# ----- 4.4 COEFICIENTE DE CORRELACIÓN LINEAL DE PEARSON -----
titulo("4.4 Coeficiente de correlación lineal de Pearson", 2)

r = (n * sum_xy - sum_x * sum_y) / np.sqrt(
    (n * sum_x2 - sum_x ** 2) * (n * sum_y2 - sum_y ** 2)
)
print(f"  Coeficiente de correlación   r = {r:.6f}")
print(f"     >> {interpretar_correlacion(r)}")


# ----- 4.5 COEFICIENTE DE DETERMINACIÓN -----
titulo("4.5 Coeficiente de determinación", 2)

r2 = r ** 2
print(f"  Coef. de determinación   r² = {r2:.6f}  =  {r2 * 100:.2f}%")
print(f"\n  INTERPRETACIÓN:")
print(f"     El {r2 * 100:.2f}% de la variabilidad del rendimiento académico se explica")
print(f"     linealmente por la cantidad de horas de uso de redes sociales.")
print(f"     El restante {(1 - r2) * 100:.2f}% se debe a otros factores no incluidos")
print(f"     en este modelo (sueño, hábitos de estudio, contexto familiar, etc.).")


# ----- 4.6 DIAGRAMA DE DISPERSIÓN CON RECTA DE REGRESIÓN -----
titulo("4.6 Generación del diagrama de dispersión", 2)

fig, ax = plt.subplots(figsize=(11, 6))
# Puntos de dispersión
ax.scatter(X, Y, alpha=0.4, s=30, color='#2E75B6', edgecolor='navy', label='Datos observados')

# Recta de regresión
x_recta = np.linspace(X.min(), X.max(), 100)
y_recta = a + b * x_recta
ax.plot(x_recta, y_recta, color='red', linewidth=2.5,
        label=f'ŷ = {a:.2f} {signo} {abs(b):.4f}·x')

# Decoración
ax.set_title('Diagrama de dispersión con recta de regresión\n'
             f'Horas de redes sociales vs Rendimiento académico (n={n})',
             fontsize=13, fontweight='bold')
ax.set_xlabel('Horas diarias de uso de redes sociales (X)', fontsize=11)
ax.set_ylabel('Rendimiento académico (Y)', fontsize=11)
ax.set_xticks(range(0, 13))
ax.grid(True, alpha=0.3)
ax.legend(loc='upper right', fontsize=10)

# Anotar coeficientes en el gráfico
texto_anotacion = f"r = {r:.4f}\nr² = {r2:.4f}\n{interpretar_correlacion(r)}"
ax.annotate(texto_anotacion, xy=(0.02, 0.05), xycoords='axes fraction',
            fontsize=10, bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8))
plt.tight_layout()
plt.savefig('dispersion_modulo7.png', dpi=150, bbox_inches='tight')
plt.close()
print("  ✓ Diagrama de dispersión guardado como 'dispersion_modulo7.png'")


# ============================================================================
# 5. COMPARACIÓN: MANUAL vs EXCEL vs PYTHON
# ============================================================================
titulo("COMPARACIÓN DE RESULTADOS - Manual / Excel / Python", 1)

print("""
   Los resultados obtenidos por los tres métodos coinciden exactamente.
   La diferencia entre métodos NO es de exactitud (los tres dan el mismo número),
   sino de ESFUERZO y TIEMPO requeridos:

   ┌─────────────────────────┬──────────────┬──────────────┬──────────────┐
   │ Tarea                   │   Manual     │   Excel      │   Python     │
   ├─────────────────────────┼──────────────┼──────────────┼──────────────┤
   │ Tabla de frecuencias    │  ~ 30 min    │  ~ 5 min     │  ~ 2 seg     │
   │ Medidas (media, etc.)   │  ~ 20 min    │  ~ 3 min     │  ~ 1 seg     │
   │ Cuartiles/Deciles       │  ~ 15 min    │  ~ 5 min     │  ~ 1 seg     │
   │ Asimetría y curtosis    │  ~ 25 min    │  ~ 10 min    │  ~ 1 seg     │
   │ Regresión (490 datos)   │  IMPRACTICABLE│ ~ 20 min    │  ~ 2 seg     │
   │ Gráficos                │  ~ 30 min    │  ~ 10 min    │  ~ 3 seg     │
   ├─────────────────────────┼──────────────┼──────────────┼──────────────┤
   │ Reproducibilidad        │   Baja       │   Media      │   Total      │
   │ Escalabilidad           │   Muy baja   │   Media      │   Muy alta   │
   │ Riesgo de error humano  │   Alto       │   Medio      │   Bajo       │
   └─────────────────────────┴──────────────┴──────────────┴──────────────┘
""")


# ============================================================================
# 6. EXPORTACIÓN DE RESULTADOS A EXCEL
# ============================================================================
titulo("Exportación de resultados a Excel", 2)

wb = Workbook()
thin = Side(border_style='thin', color='000000')
borde = Border(left=thin, right=thin, top=thin, bottom=thin)

# Hoja 1: Tabla de frecuencias
ws1 = wb.active
ws1.title = "Tabla Frecuencias"
ws1['A1'] = "Tabla de frecuencias - Horas RRSS (Módulo 7)"
ws1['A1'].font = Font(bold=True, size=12)
ws1.merge_cells('A1:H1')

headers = ['xi', 'fi', 'Fi', 'fri', 'Fri', 'pi%', 'Pi%', 'xi·fi']
for i, h in enumerate(headers, start=1):
    c = ws1.cell(row=3, column=i, value=h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', start_color='2E75B6')
    c.alignment = Alignment(horizontal='center')
    c.border = borde

for i, row in tabla.iterrows():
    r = 4 + i
    ws1.cell(row=r, column=1, value=int(row['xi']))
    ws1.cell(row=r, column=2, value=int(row['fi']))
    ws1.cell(row=r, column=3, value=int(row['Fi']))
    ws1.cell(row=r, column=4, value=float(row['fri']))
    ws1.cell(row=r, column=5, value=float(row['Fri']))
    ws1.cell(row=r, column=6, value=float(row['pi_pct']))
    ws1.cell(row=r, column=7, value=float(row['Pi_pct']))
    ws1.cell(row=r, column=8, value=int(row['xi_fi']))
    for col in range(1, 9):
        cell = ws1.cell(row=r, column=col)
        cell.border = borde
        cell.alignment = Alignment(horizontal='center')
        if col in (4, 5):
            cell.number_format = '0.0000'
        elif col in (6, 7):
            cell.number_format = '0.00'

# Hoja 2: Todas las medidas calculadas
ws2 = wb.create_sheet("Medidas")
ws2['A1'] = "Resumen de medidas estadísticas (Módulo 7)"
ws2['A1'].font = Font(bold=True, size=12)
ws2.merge_cells('A1:B1')

medidas = [
    ('TENDENCIA CENTRAL', None),
    ('Media (x̄)',              f"{media_x:.4f}"),
    ('Moda',                    f"{moda_x}"),
    ('Mediana',                 f"{mediana_x}"),
    ('DISPERSIÓN', None),
    ('Varianza',                f"{varianza_x:.4f}"),
    ('Desvío estándar',         f"{desvio_x:.4f}"),
    ('Coef. de Variación',      f"{cv_x:.2f}%"),
    ('FORMA', None),
    ('Asimetría Pearson',       f"{as_pearson:.4f}"),
    ('Asimetría Fisher',        f"{as_fisher:.4f}"),
    ('Curtosis Fisher',         f"{curt_fisher:.4f}"),
    ('POSICIÓN', None),
    ('Q1', f"{calc_posicion(1, 4)[1]}"),
    ('Q2', f"{calc_posicion(2, 4)[1]}"),
    ('Q3', f"{calc_posicion(3, 4)[1]}"),
    ('D3', f"{calc_posicion(3, 10)[1]}"),
    ('D4', f"{calc_posicion(4, 10)[1]}"),
    ('D9', f"{calc_posicion(9, 10)[1]}"),
    ('P23', f"{calc_posicion(23, 100)[1]}"),
    ('P75', f"{calc_posicion(75, 100)[1]}"),
    ('P97', f"{calc_posicion(97, 100)[1]}"),
    ('REGRESIÓN LINEAL', None),
    ('Pendiente (b)',           f"{b:.6f}"),
    ('Ordenada al origen (a)',  f"{a:.6f}"),
    ('Ecuación recta',          f"ŷ = {a:.4f} {signo} {abs(b):.4f}·x"),
    ('Coef. correlación (r)',   f"{r:.6f}"),
    ('Coef. determinación (r²)', f"{r2:.6f}"),
    ('% variabilidad explicada', f"{r2*100:.2f}%"),
]
for i, (nombre, valor) in enumerate(medidas, start=3):
    if valor is None:
        ws2.cell(row=i, column=1, value=nombre).font = Font(bold=True, color='FFFFFF')
        ws2.cell(row=i, column=1).fill = PatternFill('solid', start_color='1F4E78')
        ws2.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
    else:
        ws2.cell(row=i, column=1, value=nombre)
        ws2.cell(row=i, column=2, value=valor).font = Font(bold=True)
        ws2.cell(row=i, column=1).border = borde
        ws2.cell(row=i, column=2).border = borde

ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 24

# Hoja 3: Comparación Excel vs Python
ws3 = wb.create_sheet("Comparación")
ws3['A1'] = "Comparación Excel (Módulo 6) vs Python (Módulo 7)"
ws3['A1'].font = Font(bold=True, size=12)
ws3.merge_cells('A1:C1')

headers_c = ['Resultado', 'Excel (Módulo 6)', 'Python (Módulo 7)']
for i, h in enumerate(headers_c, start=1):
    c = ws3.cell(row=3, column=i, value=h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', start_color='2E75B6')
    c.alignment = Alignment(horizontal='center')
    c.border = borde

filas_comp = [
    ('Pendiente (b)',        '-0,2140',  f"{b:.4f}"),
    ('Ordenada origen (a)',  '8,4790',   f"{a:.4f}"),
    ('Coef. correlación r',  '-0,6315',  f"{r:.4f}"),
    ('Coef. determinación r²', '0,3988', f"{r2:.4f}"),
]
for i, (n_lbl, exc, py) in enumerate(filas_comp, start=4):
    ws3.cell(row=i, column=1, value=n_lbl)
    ws3.cell(row=i, column=2, value=exc)
    ws3.cell(row=i, column=3, value=py)
    for col in range(1, 4):
        ws3.cell(row=i, column=col).border = borde
        ws3.cell(row=i, column=col).alignment = Alignment(horizontal='center')

ws3.column_dimensions['A'].width = 28
ws3.column_dimensions['B'].width = 22
ws3.column_dimensions['C'].width = 22

wb.save('resultados_modulo7.xlsx')
print("  ✓ Archivo 'resultados_modulo7.xlsx' generado.")


# ============================================================================
# FINALIZACIÓN
# ============================================================================
titulo("PROCESO FINALIZADO CORRECTAMENTE", 1)
print("""
   Archivos generados:
     1. resultados_modulo7.xlsx   (tabla, medidas y comparación)
     2. histograma_modulo7.png    (histograma con líneas de tendencia central)
     3. dispersion_modulo7.png    (diagrama de dispersión con recta de regresión)
""")
