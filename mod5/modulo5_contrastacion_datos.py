"""
================================================================================
TRABAJO PRÁCTICO INTEGRADOR - ESTADÍSTICA
MÓDULO 5 - CUARTA ETAPA: CONTRASTACIÓN DE DATOS
================================================================================
Análisis estadístico descriptivo sobre los datos proporcionados por la cátedra
(490 estudiantes) y comparación con los datos de la primera etapa (200 estudiantes
encuestados por nuestro grupo).

Variable de análisis: Cantidad de horas diarias de consumo de redes sociales.

Librerías utilizadas:
- pandas: manejo de datos en estructuras tipo tabla (DataFrame)
- numpy:  cálculos numéricos y estadísticos
- matplotlib: generación de gráficos (histograma)
- openpyxl: exportación de resultados a Excel
================================================================================
"""

# ============================================================================
# 1. IMPORTACIÓN DE LIBRERÍAS
# ============================================================================
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ============================================================================
# 2. CARGA DE DATOS DEL EXCEL DE LA CÁTEDRA
# ============================================================================
print("=" * 80)
print("MÓDULO 5 - CUARTA ETAPA: CONTRASTACIÓN DE DATOS")
print("=" * 80)

# Ruta del archivo provisto por la docente
RUTA_EXCEL = "tp_proporcionado.xlsx"

# Cargar la hoja 'Datos Brutos'
df = pd.read_excel(RUTA_EXCEL, sheet_name='Datos Brutos')

# Renombrar columnas para trabajar más cómodos (la original tiene dos espacios)
df.columns = ['Horas_RRSS', 'Rendimiento']

# Quedarnos solo con la variable solicitada
horas = df['Horas_RRSS']
n = len(horas)
print(f"\n>> Datos cargados correctamente.")
print(f"   Total de registros (n): {n}")
print(f"   Variable de análisis: Horas diarias de consumo de redes sociales\n")


# ============================================================================
# 3. CONSTRUCCIÓN DE LA TABLA DE FRECUENCIAS
# ============================================================================
print("=" * 80)
print("3. TABLA DE FRECUENCIAS")
print("=" * 80)

# Frecuencia absoluta (fi): cuántas veces aparece cada valor
fi = horas.value_counts().sort_index()

# Crear el DataFrame de la tabla
tabla = pd.DataFrame({'xi': fi.index, 'fi': fi.values})

# Frecuencia absoluta acumulada (Fi)
tabla['Fi'] = tabla['fi'].cumsum()

# Frecuencia relativa (fri = fi/n)
tabla['fri'] = tabla['fi'] / n

# Frecuencia relativa acumulada (Fri)
tabla['Fri'] = tabla['fri'].cumsum()

# Porcentajes
tabla['pi_pct'] = tabla['fri'] * 100
tabla['Pi_pct'] = tabla['Fri'] * 100

# Columnas auxiliares para cálculos posteriores (media, varianza, asimetría, curtosis)
tabla['xi_fi'] = tabla['xi'] * tabla['fi']

# Mostrar tabla
print(tabla.to_string(index=False, float_format='%.4f'))
print(f"\n   TOTAL: n = {tabla['fi'].sum()}")
print(f"   Suma fri = {tabla['fri'].sum():.4f}")
print(f"   Suma pi_pct = {tabla['pi_pct'].sum():.2f}%")


# ============================================================================
# 4. MEDIDAS DE TENDENCIA CENTRAL
# ============================================================================
print("\n" + "=" * 80)
print("4. MEDIDAS DE TENDENCIA CENTRAL")
print("=" * 80)

# Media aritmética: x̄ = Σ(xi · fi) / n
media = tabla['xi_fi'].sum() / n

# Moda: el xi con mayor fi
moda = tabla.loc[tabla['fi'].idxmax(), 'xi']

# Mediana: valor en la posición n/2 según Fi
pos_mediana = n / 2
mediana = tabla.loc[tabla['Fi'] >= pos_mediana, 'xi'].iloc[0]

print(f"   Media aritmética (x̄) = {media:.4f} horas/día")
print(f"   Moda (Mo)            = {moda} horas/día")
print(f"   Mediana (Md)         = {mediana} horas/día")


# ============================================================================
# 5. MEDIDAS DE DISPERSIÓN
# ============================================================================
print("\n" + "=" * 80)
print("5. MEDIDAS DE DISPERSIÓN")
print("=" * 80)

# Agregar columnas auxiliares al DataFrame
tabla['dif2_fi'] = (tabla['xi'] - media) ** 2 * tabla['fi']
tabla['dif3_fi'] = (tabla['xi'] - media) ** 3 * tabla['fi']
tabla['dif4_fi'] = (tabla['xi'] - media) ** 4 * tabla['fi']

# Varianza muestral (dividida por n-1, como se trabajó en clase)
varianza = tabla['dif2_fi'].sum() / (n - 1)

# Desvío estándar
desvio = np.sqrt(varianza)

# Coeficiente de variación
cv = (desvio / media) * 100

print(f"   Σ(xi - x̄)² · fi = {tabla['dif2_fi'].sum():.4f}")
print(f"   Varianza (s²)    = {varianza:.4f}")
print(f"   Desvío (s)       = {desvio:.4f}")
print(f"   Coef. Variación  = {cv:.2f}%   →  {'HETEROGÉNEOS (>30%)' if cv > 30 else 'HOMOGÉNEOS'}")


# ============================================================================
# 6. MEDIDAS DE FORMA (ASIMETRÍA Y CURTOSIS)
# ============================================================================
print("\n" + "=" * 80)
print("6. MEDIDAS DE FORMA: ASIMETRÍA Y CURTOSIS")
print("=" * 80)

# Asimetría de Pearson: As = 3(x̄ - Md)/s
as_pearson = 3 * (media - mediana) / desvio

# Asimetría de Fisher (datos agrupados): As = Σf(x-x̄)³ / (n · s³)
as_fisher = tabla['dif3_fi'].sum() / (n * desvio ** 3)

# Curtosis de Fisher (datos agrupados): α = Σf(x-x̄)⁴ / (n · s⁴)
curt_fisher = tabla['dif4_fi'].sum() / (n * desvio ** 4)

# Interpretación textual
def interp_asimetria(v):
    if v < -0.1:
        return "Asimétrica negativa (cola a la izquierda)"
    elif v > 0.1:
        return "Asimétrica positiva (cola a la derecha)"
    else:
        return "Aproximadamente simétrica"

def interp_curtosis(v):
    if v < 3:
        return f"Platicúrtica (α < 3, distribución achatada)"
    elif v > 3:
        return f"Leptocúrtica (α > 3, distribución apuntada)"
    else:
        return "Mesocúrtica (α = 3, como la curva normal)"

print(f"   Asimetría de Pearson  = {as_pearson:.4f}")
print(f"      → {interp_asimetria(as_pearson)}")
print(f"   Asimetría de Fisher   = {as_fisher:.4f}")
print(f"      → {interp_asimetria(as_fisher)}")
print(f"   Curtosis de Fisher    = {curt_fisher:.4f}")
print(f"      → {interp_curtosis(curt_fisher)}")


# ============================================================================
# 7. MEDIDAS DE POSICIÓN: CUARTILES, DECILES, PERCENTILES
# ============================================================================
print("\n" + "=" * 80)
print("7. MEDIDAS DE POSICIÓN (método por posición sobre Fi)")
print("=" * 80)

def calcular_posicion(k, total, base):
    """Calcula la posición k/base sobre el total y devuelve el xi correspondiente"""
    posicion = k * total / base
    valor = tabla.loc[tabla['Fi'] >= posicion, 'xi'].iloc[0]
    return posicion, valor

print("\n   CUARTILES:")
for k in [1, 2, 3]:
    pos, val = calcular_posicion(k, n, 4)
    print(f"   Q{k}: posición = {k}·{n}/4 = {pos:.0f}  →  Q{k} = {val} horas")

print("\n   DECILES:")
for k in [3, 4, 9]:
    pos, val = calcular_posicion(k, n, 10)
    print(f"   D{k}: posición = {k}·{n}/10 = {pos:.0f}  →  D{k} = {val} horas")

print("\n   PERCENTILES:")
for k in [23, 75, 97]:
    pos, val = calcular_posicion(k, n, 100)
    print(f"   P{k}: posición = {k}·{n}/100 = {pos:.0f}  →  P{k} = {val} horas")


# ============================================================================
# 8. COMPARACIÓN CON PRIMERA ETAPA (200 estudiantes encuestados por el grupo)
# ============================================================================
print("\n" + "=" * 80)
print("8. COMPARACIÓN CON LA PRIMERA ETAPA")
print("=" * 80)

# Resultados de nuestra encuesta (Módulo 3)
primera_etapa = {
    'n': 200,
    'media': 4.88,
    'mediana': 5,
    'moda': 5,
    'varianza': 7.8046,
    'desvio': 2.7937,
    'cv': 57.25,
    'as_pearson': -0.1289,
    'as_fisher': 0.4117,
    'curt_fisher': 2.5889,
}

# Resultados de la cuarta etapa (490 datos del Excel docente)
cuarta_etapa = {
    'n': n,
    'media': media,
    'mediana': mediana,
    'moda': moda,
    'varianza': varianza,
    'desvio': desvio,
    'cv': cv,
    'as_pearson': as_pearson,
    'as_fisher': as_fisher,
    'curt_fisher': curt_fisher,
}

print(f"\n   {'Medida':<25} {'1ra Etapa (n=200)':>20} {'4ta Etapa (n=490)':>20} {'Diferencia':>14}")
print(f"   {'-'*25} {'-'*20} {'-'*20} {'-'*14}")
labels = [
    ('Media aritmética',     'media',       '{:.4f}'),
    ('Mediana',              'mediana',     '{:.0f}'),
    ('Moda',                 'moda',        '{:.0f}'),
    ('Varianza',             'varianza',    '{:.4f}'),
    ('Desvío estándar',      'desvio',      '{:.4f}'),
    ('Coef. de Variación %', 'cv',          '{:.2f}'),
    ('Asimetría Pearson',    'as_pearson',  '{:.4f}'),
    ('Asimetría Fisher',     'as_fisher',   '{:.4f}'),
    ('Curtosis Fisher',      'curt_fisher', '{:.4f}'),
]
for nombre, key, fmt in labels:
    v1 = primera_etapa[key]
    v2 = cuarta_etapa[key]
    dif = v2 - v1
    print(f"   {nombre:<25} {fmt.format(v1):>20} {fmt.format(v2):>20} {dif:>14.4f}")

# Cálculo de la medida pertinente de comparación: el CV
print("\n" + "-" * 80)
print("   MEDIDA PERTINENTE PARA COMPARAR LAS DOS MUESTRAS: COEFICIENTE DE VARIACIÓN")
print("-" * 80)
print(f"   CV (1ra etapa, n=200): {primera_etapa['cv']:.2f}%")
print(f"   CV (4ta etapa, n=490): {cuarta_etapa['cv']:.2f}%")
print(f"   Diferencia:            {cuarta_etapa['cv'] - primera_etapa['cv']:.2f} puntos porcentuales")
print(f"""
   INTERPRETACIÓN:
   El Coeficiente de Variación es la medida adecuada para comparar la dispersión
   relativa de dos muestras de tamaños distintos (200 vs 490), ya que está
   expresado como porcentaje y es independiente de la unidad de medida.
   
   En la primera etapa, el CV es {primera_etapa['cv']:.2f}%, mientras que en la cuarta etapa
   es {cuarta_etapa['cv']:.2f}%. Ambos valores superan el umbral del 30%, lo que confirma que
   los datos son HETEROGÉNEOS en ambos casos.
   
   La muestra de nuestra encuesta (200 estudiantes) presenta una dispersión 
   relativa MAYOR ({primera_etapa['cv']:.2f}% vs {cuarta_etapa['cv']:.2f}%), lo que indica que
   nuestros datos están más alejados de su media que los del Excel de la cátedra.
""")


# ============================================================================
# 9. HISTOGRAMA
# ============================================================================
print("=" * 80)
print("9. GENERACIÓN DEL HISTOGRAMA")
print("=" * 80)

fig, ax = plt.subplots(figsize=(11, 6))
ax.bar(tabla['xi'], tabla['fi'], width=1.0, edgecolor='black',
       color='#2E75B6', alpha=0.85)

# Líneas verticales: media, mediana, moda
ax.axvline(media, color='red', linestyle='--', linewidth=2,
           label=f'Media = {media:.2f}')
ax.axvline(mediana, color='green', linestyle=':', linewidth=2,
           label=f'Mediana = {mediana}')
ax.axvline(moda, color='orange', linestyle='-.', linewidth=2,
           label=f'Moda = {moda}')

ax.set_title('Histograma - Cantidad de horas diarias de consumo de redes sociales\n'
             f'(n = {n} estudiantes - datos cátedra)',
             fontsize=13, fontweight='bold')
ax.set_xlabel('Horas diarias de uso (xi)', fontsize=11)
ax.set_ylabel('Cantidad de estudiantes (fi)', fontsize=11)
ax.set_xticks(range(0, 13))
ax.grid(axis='y', alpha=0.3)
ax.legend(loc='upper right', fontsize=10)

# Etiquetas con frecuencia encima de cada barra
for x, f in zip(tabla['xi'], tabla['fi']):
    ax.text(x, f + 1, str(f), ha='center', fontsize=9)

plt.tight_layout()
plt.savefig('histograma_modulo5.png', dpi=150, bbox_inches='tight')
print("   ✓ Histograma guardado como 'histograma_modulo5.png'")


# ============================================================================
# 10. EXPORTACIÓN DE RESULTADOS A EXCEL
# ============================================================================
print("\n" + "=" * 80)
print("10. EXPORTACIÓN A EXCEL")
print("=" * 80)

wb = Workbook()
thin = Side(border_style='thin', color='000000')
borde = Border(left=thin, right=thin, top=thin, bottom=thin)

# ---- Hoja 1: Tabla de frecuencias
ws1 = wb.active
ws1.title = "Tabla Frecuencias"
ws1['A1'] = "Tabla de frecuencias - Horas diarias de uso de redes sociales (datos cátedra)"
ws1['A1'].font = Font(bold=True, size=12)
ws1.merge_cells('A1:H1')

headers = ['xi', 'fi', 'Fi', 'fri', 'Fri', 'pi%', 'Pi%', 'xi·fi']
for i, h in enumerate(headers, start=1):
    c = ws1.cell(row=3, column=i, value=h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', start_color='2E75B6')
    c.alignment = Alignment(horizontal='center')
    c.border = borde

for i, row in tabla.iterrows():
    r = 4 + i
    ws1.cell(row=r, column=1, value=int(row['xi']))
    ws1.cell(row=r, column=2, value=int(row['fi']))
    ws1.cell(row=r, column=3, value=int(row['Fi']))
    ws1.cell(row=r, column=4, value=float(row['fri']))
    ws1.cell(row=r, column=5, value=float(row['Fri']))
    ws1.cell(row=r, column=6, value=float(row['pi_pct']))
    ws1.cell(row=r, column=7, value=float(row['Pi_pct']))
    ws1.cell(row=r, column=8, value=int(row['xi_fi']))
    for col in range(1, 9):
        cell = ws1.cell(row=r, column=col)
        cell.border = borde
        cell.alignment = Alignment(horizontal='center')
        if col in (4, 5):
            cell.number_format = '0.0000'
        elif col in (6, 7):
            cell.number_format = '0.00'

# Fila TOTAL
tot_row = 4 + len(tabla)
ws1.cell(row=tot_row, column=1, value="TOTAL").font = Font(bold=True)
ws1.cell(row=tot_row, column=2, value=int(tabla['fi'].sum())).font = Font(bold=True)
ws1.cell(row=tot_row, column=4, value=1.0).font = Font(bold=True)
ws1.cell(row=tot_row, column=6, value=100.0).font = Font(bold=True)
ws1.cell(row=tot_row, column=8, value=int(tabla['xi_fi'].sum())).font = Font(bold=True)
for col in range(1, 9):
    ws1.cell(row=tot_row, column=col).fill = PatternFill('solid', start_color='FFE699')
    ws1.cell(row=tot_row, column=col).border = borde
    ws1.cell(row=tot_row, column=col).alignment = Alignment(horizontal='center')

for col, w in [('A',8),('B',8),('C',8),('D',10),('E',10),('F',10),('G',10),('H',12)]:
    ws1.column_dimensions[col].width = w

# ---- Hoja 2: Medidas calculadas
ws2 = wb.create_sheet("Medidas Calculadas")
ws2['A1'] = "Resumen de medidas estadísticas - Cuarta Etapa (n=490)"
ws2['A1'].font = Font(bold=True, size=12)
ws2.merge_cells('A1:C1')

medidas_lista = [
    ('TENDENCIA CENTRAL', None),
    ('Media aritmética',      f"{media:.4f}"),
    ('Moda',                  f"{moda}"),
    ('Mediana',               f"{mediana}"),
    ('DISPERSIÓN', None),
    ('Varianza muestral',     f"{varianza:.4f}"),
    ('Desvío estándar',       f"{desvio:.4f}"),
    ('Coef. de Variación',    f"{cv:.2f}%"),
    ('FORMA', None),
    ('Asimetría de Pearson',  f"{as_pearson:.4f}"),
    ('Asimetría de Fisher',   f"{as_fisher:.4f}"),
    ('Curtosis de Fisher',    f"{curt_fisher:.4f}"),
    ('POSICIÓN', None),
    ('Q1', f"{calcular_posicion(1, n, 4)[1]}"),
    ('Q2', f"{calcular_posicion(2, n, 4)[1]}"),
    ('Q3', f"{calcular_posicion(3, n, 4)[1]}"),
    ('D3', f"{calcular_posicion(3, n, 10)[1]}"),
    ('D4', f"{calcular_posicion(4, n, 10)[1]}"),
    ('D9', f"{calcular_posicion(9, n, 10)[1]}"),
    ('P23', f"{calcular_posicion(23, n, 100)[1]}"),
    ('P75', f"{calcular_posicion(75, n, 100)[1]}"),
    ('P97', f"{calcular_posicion(97, n, 100)[1]}"),
]
for i, (nombre, valor) in enumerate(medidas_lista, start=3):
    if valor is None:
        ws2.cell(row=i, column=1, value=nombre).font = Font(bold=True, color='FFFFFF')
        ws2.cell(row=i, column=1).fill = PatternFill('solid', start_color='1F4E78')
        ws2.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
    else:
        ws2.cell(row=i, column=1, value=nombre)
        ws2.cell(row=i, column=2, value=valor).font = Font(bold=True)
        ws2.cell(row=i, column=1).border = borde
        ws2.cell(row=i, column=2).border = borde

ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 18

# ---- Hoja 3: Comparación con primera etapa
ws3 = wb.create_sheet("Comparación 1ra vs 4ta")
ws3['A1'] = "Comparación: Primera Etapa (n=200) vs Cuarta Etapa (n=490)"
ws3['A1'].font = Font(bold=True, size=12)
ws3.merge_cells('A1:D1')

headers3 = ['Medida', '1ra Etapa (n=200)', '4ta Etapa (n=490)', 'Diferencia']
for i, h in enumerate(headers3, start=1):
    c = ws3.cell(row=3, column=i, value=h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', start_color='2E75B6')
    c.alignment = Alignment(horizontal='center')
    c.border = borde

for i, (nombre, key, _) in enumerate(labels, start=4):
    v1 = primera_etapa[key]
    v2 = cuarta_etapa[key]
    ws3.cell(row=i, column=1, value=nombre)
    ws3.cell(row=i, column=2, value=round(v1, 4))
    ws3.cell(row=i, column=3, value=round(v2, 4))
    ws3.cell(row=i, column=4, value=round(v2 - v1, 4))
    for col in range(1, 5):
        ws3.cell(row=i, column=col).border = borde
        ws3.cell(row=i, column=col).alignment = Alignment(horizontal='center')

ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 22
ws3.column_dimensions['C'].width = 22
ws3.column_dimensions['D'].width = 14

# Guardar
wb.save('resultados_modulo5.xlsx')
print("   ✓ Resultados guardados en 'resultados_modulo5.xlsx'")

print("\n" + "=" * 80)
print("PROCESO FINALIZADO CORRECTAMENTE")
print("Archivos generados:")
print("  1. resultados_modulo5.xlsx  (tabla + medidas + comparación)")
print("  2. histograma_modulo5.png   (histograma con media, mediana y moda)")
print("=" * 80)
